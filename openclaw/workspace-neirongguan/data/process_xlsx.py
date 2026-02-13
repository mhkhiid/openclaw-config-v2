#!/usr/bin/env python3
"""
将历史媒体数据存档xlsx转换为标准化CSV
按 data_ingestion.md 规则做分类
"""

import openpyxl
import csv
import re
from datetime import datetime

# ============ Classification Rules ============

def classify_topic(text):
    """从钩子/脚本/标题推断 topic_category"""
    if not text:
        return "其他", ""
    t = text[:500]  # 主要看前500字
    
    # 按优先级匹配
    if re.search(r'加班|违法|工时|超时|不敢加班|强制休息|不让加班|催.*下班', t):
        cat = "权益加班"
    elif re.search(r'夜班|大夜|值班|夜班费|夜班补贴|专职夜班', t) and not re.search(r'年假|病假|请假|休假', t[:50]):
        cat = "夜班"
    elif re.search(r'年假|病假|请假|休假|带薪假|排假|休息日', t):
        cat = "休假病假"
    elif re.search(r'护士证|出国|认证|德语|B1|签证|中专逆袭|来德国.*路径|隐藏价值', t):
        cat = "护士证路径"
    elif re.search(r'为什么不回国|来德国\d+年|我的经历|改变命运|不是不想回|回不去', t):
        cat = "个人故事"
    elif re.search(r'床护比|几个护士管|排班|人员配置|互换比|28张.*床|21个.*护士', t):
        cat = "互换比管理"
    elif re.search(r'工资|收入|到手|税后|涨薪|小费|13薪|起薪|月入|欧.*薪|薪.*欧', t):
        cat = "工资收入"
    elif re.search(r'国内.*德国.*对比|德国.*国内.*区别|中德.*差异|差异', t) and t.count('一') + t.count('二') + t.count('三') >= 2:
        cat = "中德差异"
    elif re.search(r'护士长|护理部', t) and not re.search(r'违法|加班|请假', t[:50]):
        cat = "护士长"
    elif re.search(r'学员|招聘|公司介绍|集体合影|海狸未来|赴德|到达.*法兰克福|海丽学员|面试', t):
        cat = "招聘学员"
    elif re.search(r'压疮|护理不良|护患关系|医患', t):
        cat = "中德差异"
    else:
        cat = "其他"
    
    # topic_angle: 5-10字概括
    angle = extract_angle(t, cat)
    return cat, angle

def extract_angle(text, cat):
    """从内容提取5-10字的切入点"""
    t = text[:100]
    
    # 尝试从开头提取关键词
    patterns = {
        "权益加班": [r'(加班违法)', r'(不敢加班)', r'(催.*下班)', r'(强制休息)', r'(工时规定)'],
        "夜班": [r'(夜班.*抢)', r'(夜班.*收入)', r'(夜班.*流程)', r'(专职夜班)', r'(夜班补贴)'],
        "休假病假": [r'(年假\d+)', r'(病假)', r'(全科请假)', r'(排.*年假)', r'(带薪假)'],
        "护士证路径": [r'(护士证.*价值)', r'(出国路径)', r'(中专.*逆袭)'],
        "个人故事": [r'(为什么不回国)', r'(不是不想回)', r'(来德国\d+年)'],
        "工资收入": [r'(起薪)', r'(工资)', r'(小费)', r'(收入)', r'(涨薪)'],
        "招聘学员": [r'(学员.*到达)', r'(赴德)', r'(面试)'],
    }
    
    for p in patterns.get(cat, []):
        m = re.search(p, t)
        if m:
            return m.group(1)[:10]
    
    # fallback: 取前10个有意义的字
    clean = re.sub(r'[#@\s]', '', t[:20])
    return clean[:10] if clean else "—"

def classify_hook_model(hook_text):
    """从钩子内容判断 hook_model"""
    if not hook_text:
        return "无明确模型"
    h = hook_text
    
    # A_反常识
    if re.search(r'居然|竟然|你敢信|没想到|真服了|违法|不敢|谁能想到', h):
        return "A_反常识"
    
    # B_数字冲击
    nums = re.findall(r'\d+', h)
    big_nums = [n for n in nums if int(n) >= 30 or (int(n) >= 2 and '万' in h) or (int(n) >= 1000)]
    if big_nums and re.search(r'块|欧|万|天|小时|星期|个月|岁', h):
        return "B_数字冲击"
    
    # C_共情痛点
    if re.search(r'中专|找不到工作|丫鬟证|为什么不回|不想回|回不去|学历|改变命运', h):
        return "C_共情痛点"
    
    # D_冲突事件
    if re.search(r'被.*骂|请假|爆炸|崩溃|出事|投诉|吵架|只剩', h):
        return "D_冲突事件"
    
    # E_三连否定
    if h.count('不是') >= 2 and '而是' in h:
        return "E_三连否定"
    
    # F_假设反推
    if re.search(r'如果|假如', h) and re.search(r'搬到国内|放到中国|在国内', h):
        return "F_假设反推"
    
    # 有套路但不明确
    if re.search(r'你知道|你见过|你猜|你以为', h):
        return "其他"
    
    # 平淡开头
    if re.search(r'^(和大家|今天来|大家好|我发现)', h):
        return "无明确模型"
    
    return "其他"

def extract_hook(script, max_chars=60):
    """从脚本提取前5秒的钩子文本（约前50-60字）"""
    if not script:
        return ""
    clean = script.strip().replace('\n', ' ').replace('\r', '')
    # 在自然断句处截断
    if len(clean) <= max_chars:
        return clean
    # 找最近的断句点
    cutoff = clean[:max_chars]
    for sep in [' ', '，', '。', '！', '？', '了', '的', '吗', '嘞', '呢']:
        idx = cutoff.rfind(sep)
        if idx > 30:
            return cutoff[:idx+1].strip()
    return cutoff.strip()

def calc_performance_tier(views):
    """播放量等级"""
    if views is None:
        return "—"
    try:
        v = int(views)
    except:
        return "—"
    if v >= 1000000: return "S"
    if v >= 100000: return "A"
    if v >= 50000: return "B"
    if v >= 10000: return "C"
    return "D"

def safe_float(val):
    """安全转换为浮点数"""
    if val is None:
        return None
    try:
        f = float(val)
        if f < 1 and f > 0:  # 0.344 -> 34.4%
            return round(f * 100, 2)
        return round(f, 2)
    except:
        return None

def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except:
        return None

def count_chinese_chars(text):
    """计算中文字符数"""
    if not text:
        return None
    return len(re.findall(r'[\u4e00-\u9fff]', text))

def scan_pillars(text):
    """从脚本中扫描数据支柱"""
    if not text:
        return ""
    pillars = []
    checks = {
        'P1': r'28张.*床',
        'P2': r'21个.*护士|20多个护士',
        'P3': r'早班\d|晚班\d|夜班\d',
        'P4': r'38\.5.*小时|38\.5.*工时',
        'P5': r'超时.*补休|加班.*违法|加班.*补',
        'P6': r'换衣.*工时|换衣.*算',
        'P7': r'10.*小时.*夜班|夜班.*10.*小时',
        'P8': r'2000.*块|200.*欧.*夜班|夜班.*200.*欧|夜班.*2000',
        'P9': r'最多巡视.*3|巡视.*3.*趟|3.*趟',
        'P10': r'10个夜班.*1天|夜班.*多.*天.*假',
        'P11': r'30天.*年假|年假.*30',
        'P12': r'6周.*病假|病假.*6周',
        'P13': r'3天.*假条|病假.*条',
        'P14': r'3414|3400.*欧|3300.*欧|起薪',
        'P15': r'税后.*到手|到手.*2[89]00|2800.*3200',
        'P16': r'13薪|圣诞.*金|圣诞.*奖',
        'P17': r'工龄.*涨|按.*年.*涨|自动.*涨',
        'P18': r'工资.*标准.*透明|网上.*可查',
        'P19': r'永久.*合同',
        'P20': r'不加班.*不考试|不用.*考试|没有.*考试',
        'P21': r'免费医疗|免费.*教育',
        'P22': r'中专',
        'P23': r'在德国.*\d+年|做护士.*\d+年',
        'P24': r'17岁|18岁|19岁|来德国.*时',
        'P25': r'护士.*挑.*医院|你挑.*工作',
        'P26': r'中专.*大专.*本科.*同|不分.*等级|学历.*不.*分',
        'P27': r'不限.*年龄|63岁|没有.*年龄',
        'P28': r'拿.*证.*不再.*考|没有.*考试',
    }
    for pid, pattern in checks.items():
        if re.search(pattern, text):
            pillars.append(pid)
    return ','.join(pillars)

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # Try common formats
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except:
            continue
    return s[:10]

def format_time(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime('%H:%M')
    s = str(val).strip()
    try:
        return datetime.strptime(s, '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
    except:
        return ""

# ============ Process Each Sheet ============

STANDARD_HEADERS = [
    'video_id', 'account', 'publish_date', 'publish_time', 'topic_category',
    'topic_angle', 'format', 'hook_text', 'hook_model', 'duration_seconds',
    'word_count', 'views', 'retention_5s', 'completion_rate', 'like_count',
    'comment_count', 'share_count', 'favorite_count', 'like_rate', 'share_rate',
    'performance_tier', 'pillars_used', 'is_repeat', 'has_script', 'script_text',
    'video_url', 'video_title', 'is_hidden', 'notes'
]

def process_sheet1(wb, start_id):
    """抖音官号（海狸未来）"""
    ws = wb['抖音官号（海狸未来）']
    rows = []
    vid = start_id
    
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            continue
        
        script = ws.cell(row=r, column=16).value
        script = str(script).strip() if script else ""
        hook = extract_hook(script) if script else ""
        topic_cat, topic_angle = classify_topic(script if script else "")
        hook_model = classify_hook_model(hook)
        
        views = safe_int(ws.cell(row=r, column=5).value)
        likes = safe_int(ws.cell(row=r, column=7).value)
        comments = safe_int(ws.cell(row=r, column=8).value)
        shares = safe_int(ws.cell(row=r, column=10).value)
        
        completion = safe_float(ws.cell(row=r, column=11).value)
        retention_5s = safe_float(ws.cell(row=r, column=13).value)
        
        like_rate = round(likes / views * 100, 2) if likes and views and views > 0 else None
        share_rate = round(shares / views * 100, 2) if shares and views and views > 0 else None
        
        is_hidden = str(ws.cell(row=r, column=4).value) if ws.cell(row=r, column=4).value else ""
        
        row = {
            'video_id': vid,
            'account': '抖音官号',
            'publish_date': format_date(date_val),
            'publish_time': format_time(ws.cell(row=r, column=2).value),
            'topic_category': topic_cat,
            'topic_angle': topic_angle,
            'format': '口播',  # default, notes if unsure
            'hook_text': hook,
            'hook_model': hook_model,
            'duration_seconds': safe_int(ws.cell(row=r, column=3).value),
            'word_count': count_chinese_chars(script) if script else None,
            'views': views,
            'retention_5s': retention_5s,
            'completion_rate': completion,
            'like_count': likes,
            'comment_count': comments,
            'share_count': shares,
            'favorite_count': None,  # not in this sheet
            'like_rate': like_rate,
            'share_rate': share_rate,
            'performance_tier': calc_performance_tier(views),
            'pillars_used': scan_pillars(script) if script else "",
            'is_repeat': 'FALSE',
            'has_script': 'TRUE' if script else 'FALSE',
            'script_text': script,
            'video_url': '',
            'video_title': '',
            'is_hidden': is_hidden,
            'notes': '格式默认口播，需确认' if not script else ''
        }
        rows.append(row)
        vid += 1
    
    return rows, vid

def process_sheet2(wb, start_id):
    """抖音小号（米娅）"""
    ws = wb['抖音小号（海狸护理出国米娅）']
    rows = []
    vid = start_id
    
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            continue
        
        title = str(ws.cell(row=r, column=4).value or "")
        script = ws.cell(row=r, column=17).value
        script = str(script).strip() if script else ""
        
        text_for_classify = script if script else title
        hook = extract_hook(script) if script else extract_hook(title)
        topic_cat, topic_angle = classify_topic(text_for_classify)
        hook_model = classify_hook_model(hook)
        
        views = safe_int(ws.cell(row=r, column=7).value)
        likes = safe_int(ws.cell(row=r, column=9).value)
        comments = safe_int(ws.cell(row=r, column=10).value)
        shares = safe_int(ws.cell(row=r, column=11).value)
        
        completion = safe_float(ws.cell(row=r, column=12).value)
        retention_5s = safe_float(ws.cell(row=r, column=14).value)
        
        like_rate = round(likes / views * 100, 2) if likes and views and views > 0 else None
        share_rate = round(shares / views * 100, 2) if shares and views and views > 0 else None
        
        is_hidden = str(ws.cell(row=r, column=6).value) if ws.cell(row=r, column=6).value else ""
        url = str(ws.cell(row=r, column=3).value or "")
        
        row = {
            'video_id': vid,
            'account': '米娅抖音',
            'publish_date': format_date(date_val),
            'publish_time': '',
            'topic_category': topic_cat,
            'topic_angle': topic_angle,
            'format': '口播',
            'hook_text': hook,
            'hook_model': hook_model,
            'duration_seconds': safe_int(ws.cell(row=r, column=5).value),
            'word_count': count_chinese_chars(script) if script else None,
            'views': views,
            'retention_5s': retention_5s,
            'completion_rate': completion,
            'like_count': likes,
            'comment_count': comments,
            'share_count': shares,
            'favorite_count': None,
            'like_rate': like_rate,
            'share_rate': share_rate,
            'performance_tier': calc_performance_tier(views),
            'pillars_used': scan_pillars(script) if script else "",
            'is_repeat': 'FALSE',
            'has_script': 'TRUE' if script else 'FALSE',
            'script_text': script,
            'video_url': url,
            'video_title': title,
            'is_hidden': is_hidden,
            'notes': ''
        }
        rows.append(row)
        vid += 1
    
    return rows, vid

def process_sheet3(wb, start_id):
    """视频号官号"""
    ws = wb['视频号官号（海狸护理出国）']
    rows = []
    vid = start_id
    
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            continue
        
        title = str(ws.cell(row=r, column=2).value or "")
        script = ws.cell(row=r, column=16).value
        script = str(script).strip() if script else ""
        
        text_for_classify = script if script else title
        hook = extract_hook(script) if script else extract_hook(title)
        topic_cat, topic_angle = classify_topic(text_for_classify)
        hook_model = classify_hook_model(hook)
        
        views = safe_int(ws.cell(row=r, column=6).value)
        likes = safe_int(ws.cell(row=r, column=8).value)
        comments = safe_int(ws.cell(row=r, column=9).value)
        shares = safe_int(ws.cell(row=r, column=10).value)
        follows = safe_int(ws.cell(row=r, column=11).value)
        
        completion = safe_float(ws.cell(row=r, column=4).value)
        
        like_rate = round(likes / views * 100, 2) if likes and views and views > 0 else None
        share_rate = round(shares / views * 100, 2) if shares and views and views > 0 else None
        
        vid_id = str(ws.cell(row=r, column=3).value or "")
        
        # 平均播放时长 -> 估算时长
        avg_play = ws.cell(row=r, column=5).value
        duration = None
        if avg_play:
            s = str(avg_play).replace('秒', '').strip()
            try:
                duration = int(float(s))
            except:
                pass
        
        row = {
            'video_id': vid,
            'account': '视频号官号',
            'publish_date': format_date(date_val),
            'publish_time': '',
            'topic_category': topic_cat,
            'topic_angle': topic_angle,
            'format': '口播',
            'hook_text': hook,
            'hook_model': hook_model,
            'duration_seconds': duration,
            'word_count': count_chinese_chars(script) if script else None,
            'views': views,
            'retention_5s': None,
            'completion_rate': completion,
            'like_count': likes,
            'comment_count': comments,
            'share_count': shares,
            'favorite_count': None,
            'like_rate': like_rate,
            'share_rate': share_rate,
            'performance_tier': calc_performance_tier(views),
            'pillars_used': scan_pillars(script) if script else "",
            'is_repeat': 'FALSE',
            'has_script': 'TRUE' if script else 'FALSE',
            'script_text': script,
            'video_url': vid_id,
            'video_title': title,
            'is_hidden': '',
            'notes': f'涨粉:{follows}' if follows else ''
        }
        rows.append(row)
        vid += 1
    
    return rows, vid

def process_sheet5(wb, start_id):
    """视频号小分队"""
    ws = wb['视频号小分队（人工登记）']
    rows = []
    vid = start_id
    
    for r in range(4, ws.max_row + 1):  # data starts row 4
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            continue
        
        views = safe_int(ws.cell(row=r, column=3).value)
        likes = safe_int(ws.cell(row=r, column=5).value)
        comments = safe_int(ws.cell(row=r, column=6).value)
        shares = safe_int(ws.cell(row=r, column=7).value)
        follows = safe_int(ws.cell(row=r, column=8).value)
        
        like_rate = round(likes / views * 100, 2) if likes and views and views > 0 else None
        share_rate = round(shares / views * 100, 2) if shares and views and views > 0 else None
        
        url = str(ws.cell(row=r, column=2).value or "")
        
        row = {
            'video_id': vid,
            'account': '视频号小分队',
            'publish_date': format_date(date_val),
            'publish_time': '',
            'topic_category': '—',
            'topic_angle': '—',
            'format': '—',
            'hook_text': '',
            'hook_model': '—',
            'duration_seconds': None,
            'word_count': None,
            'views': views,
            'retention_5s': None,
            'completion_rate': None,
            'like_count': likes,
            'comment_count': comments,
            'share_count': shares,
            'favorite_count': None,
            'like_rate': like_rate,
            'share_rate': share_rate,
            'performance_tier': calc_performance_tier(views),
            'pillars_used': '',
            'is_repeat': 'FALSE',
            'has_script': 'FALSE',
            'script_text': '',
            'video_url': url,
            'video_title': '',
            'is_hidden': '',
            'notes': f'涨粉:{follows}; 无脚本无标题，分类需人工补充' if follows else '无脚本无标题，分类需人工补充'
        }
        rows.append(row)
        vid += 1
    
    return rows, vid

# ============ Main ============

def main():
    wb = openpyxl.load_workbook("data/历史媒体数据存档(25.12.12).xlsx", data_only=True)
    
    all_rows = []
    vid = 1
    
    print("Processing Sheet 1: 抖音官号...")
    rows1, vid = process_sheet1(wb, vid)
    all_rows.extend(rows1)
    print(f"  → {len(rows1)} rows")
    
    print("Processing Sheet 2: 米娅抖音...")
    rows2, vid = process_sheet2(wb, vid)
    all_rows.extend(rows2)
    print(f"  → {len(rows2)} rows")
    
    print("Processing Sheet 3: 视频号官号...")
    rows3, vid = process_sheet3(wb, vid)
    all_rows.extend(rows3)
    print(f"  → {len(rows3)} rows")
    
    print("Processing Sheet 5: 视频号小分队...")
    rows5, vid = process_sheet5(wb, vid)
    all_rows.extend(rows5)
    print(f"  → {len(rows5)} rows")
    
    print(f"\nTotal: {len(all_rows)} rows")
    
    # Write combined CSV (without script_text for readability)
    headers_no_script = [h for h in STANDARD_HEADERS if h != 'script_text']
    with open('data/all_videos_标准化.csv', 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=headers_no_script, extrasaction='ignore')
        writer.writeheader()
        for row in all_rows:
            writer.writerow({k: (v if v is not None else '') for k, v in row.items()})
    print("Written: data/all_videos_标准化.csv")
    
    # Write per-account CSVs
    accounts = {}
    for row in all_rows:
        acc = row['account']
        if acc not in accounts:
            accounts[acc] = []
        accounts[acc].append(row)
    
    for acc, acc_rows in accounts.items():
        fname = f"data/{acc}_标准化.csv"
        with open(fname, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=STANDARD_HEADERS, extrasaction='ignore')
            writer.writeheader()
            for row in acc_rows:
                writer.writerow({k: (v if v is not None else '') for k, v in row.items()})
        print(f"Written: {fname} ({len(acc_rows)} rows)")
    
    # Write scripts separately
    script_rows = [r for r in all_rows if r.get('script_text')]
    with open('data/all_scripts_脚本库.csv', 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['video_id', 'account', 'publish_date', 'topic_category', 'topic_angle', 'hook_text', 'script_text'])
        writer.writeheader()
        for row in script_rows:
            writer.writerow({k: row.get(k, '') for k in ['video_id', 'account', 'publish_date', 'topic_category', 'topic_angle', 'hook_text', 'script_text']})
    print(f"Written: data/all_scripts_脚本库.csv ({len(script_rows)} scripts)")
    
    # Stats
    print("\n=== 统计 ===")
    for acc, acc_rows in accounts.items():
        tiers = {}
        topics = {}
        hooks = {}
        has_script = sum(1 for r in acc_rows if r['has_script'] == 'TRUE')
        for r in acc_rows:
            t = r['performance_tier']
            tiers[t] = tiers.get(t, 0) + 1
            tc = r['topic_category']
            topics[tc] = topics.get(tc, 0) + 1
            hm = r['hook_model']
            hooks[hm] = hooks.get(hm, 0) + 1
        
        print(f"\n【{acc}】{len(acc_rows)}条 | 有脚本{has_script}条")
        print(f"  等级分布: {dict(sorted(tiers.items()))}")
        print(f"  选题分布: {dict(sorted(topics.items(), key=lambda x: -x[1]))}")
        print(f"  钩子模型: {dict(sorted(hooks.items(), key=lambda x: -x[1]))}")

if __name__ == '__main__':
    main()
